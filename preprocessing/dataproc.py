import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
import csv
from io import StringIO
import pandas as pd
import numpy.lib.recfunctions as npfunc
# Seokcheon Ju is in charge

''' 
<How to work>

<Usage> 
 Input example  :
 Output example :

'''

class Coordinate:
    def __init__(self,xcor,ycor):
        self.x=xcor
        self.y=ycor

class dataproc: 
    def __init__(self,fileroute,columnname=['x','y','value'],sheetname='',mode=0,):
        """ 
        fileroute={[list of datapath],a datapath}, columnname=['x_columnname','y_columnname','x"_columnname','y"_columnname','value_columnname'], sheet=sheet name(if different)
        if it is road or double coordinate data road=true
        mode means recall from ready made data. 0==no csv file 1==have csv file
        """
        if mode==0:
            datamanager=excelmanager(fileroute,columnname,sheetname)
            self.maindata=datamanager.getdata()
            self.datalabel=datamanager.getdatalabel()
            print (self.maindata)
           # print (self.datalabel)
            self.savedata()
        if mode==1:
            self.datalabel=[]
            with open("datalabel.csv", 'r') as csvFile:
                reader = csv.reader(csvFile)
                for row in reader:
                    if row != []:
                        self.datalabel.append(''.join(row))
                csvFile.close()
            self.maindata=[]
            for difile in os.listdir(fileroute):
               # print(difile)
                if difile.split(".")[-1]=='npy' and difile!='datalabel.csv':
                    data = np.load(difile)
                    #print(data)
                    self.maindata.append(data)
                    csvFile.close()
        
    def changecolumnname(self,columnname):
        if len(columnname):
            pass
    
    def savedata(self):
        writer = csv.writer(open("datalabel.csv", 'w'))
        for row in self.datalabel:
            writer.writerow(row)
        for datpage, datlabel in zip(self.maindata,self.datalabel):
            filename=datlabel+'.csv'
            np.save(filename,datpage)           
    def getdatalabel(self):
        return self.datalabel
    def getdata(self):
        return self.maindata

class excelmanager:
    def __init__(self,datafile,columnname,sheetname=''): 
        self.resultlist=[]
        self.datalabel=[]
        if type(datafile)==list:
            for afile in datafile:
                load_exs = load_workbook(filename=afile, data_only=True)
                print("loaded "+str(afile))

                for sheet in load_exs: #several sheets
                    self.resultlist.append(self.extractdata(sheet,columnname))
                    self.datalabel.append(sheet.title)

        else:
            load_exl = load_workbook(filename=datafile, data_only=True)
            print("loaded "+str(datafile))

            for sheet in load_exl: #several sheets
                self.resultlist.append(self.extractdata(sheet,columnname))
                self.datalabel.append(sheet.title)
        
    
    def extractdata(self,load_sheet,columnname):
        result=[]
        namelist=[]
        typelist=[]
        listtype=0
        for p,r in zip(load_sheet[2],load_sheet[1]):
            for name in columnname:
                if r.value==name:
                    if isinstance(p.value,str):
                        typelist.append(object)
                    else:
                        typelist.append(np.float64)
        dat=[]
        for name in columnname:
            for r in load_sheet[1]:
                if r.value==name:
                    namelist.append(r.value)
                    dtype=0
                    tempdat=np.array([row[r.column-1].value for row in load_sheet.iter_rows(min_row=2)],order='K')[np.newaxis]
                    tempdat=tempdat.astype([(name,typelist[listtype])])
                    tempdat=tempdat.T
                    dat.append(tempdat)
                    listtype=listtype+1
        return dat

    def getdata(self):
        return self.resultlist
    def getdatalabel(self):
        return self.datalabel